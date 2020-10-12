import requests
from bs4 import BeautifulSoup
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import re
import pathlib
import random, string
import sqlite3
import datetime
import set_mappin


def create_table(conn, c):
    # テーブルの作成
    c.execute('''CREATE TABLE teikyou_hantei(id int PRIMARY KEY, mansion_name text, address text, result text, transaction_id text, created_datetime TIMESTAMP DEFAULT (datetime(CURRENT_TIMESTAMP,'localtime')))''')

def table_isexist(conn, cur):
    cur.execute("""
        SELECT COUNT(*) FROM sqlite_master 
        WHERE TYPE='table' AND name='teikyou_hantei'
        """)
    if cur.fetchone()[0] == 0:
        return False
    return True

def randomname(n):
   randlst = [random.choice(string.ascii_letters + string.digits) for i in range(n)]
   return ''.join(randlst)

def delete(con, pk):
    """ 指定したキーのデータをDELETEする """
    cur = con.cursor()
    cur.execute('delete from teikyou_hantei where id=?', (pk,))
    con.commit()

def get_all():
    conn = sqlite3.connect('teikyou_hantei.db')
    cur = conn.cursor()
    contents = cur.execute('SELECT * FROM teikyou_hantei').fetchall()
    return contents

def get_mansion_all():
    conn = sqlite3.connect('teikyou_hantei.db')
    cur = conn.cursor()
    contents = cur.execute('SELECT mansion_name FROM teikyou_hantei').fetchall()
    return contents


def db_search(cur, mansion_name):
    contents = cur.execute('SELECT * FROM teikyou_hantei WHERE mansion_name = ?', (mansion_name,)).fetchall()
    print(contents)
    if len(contents) > 0:
        return True
    else:
        return False

def past_search(mansions, mansion):
    match = 0
    for x in mansions:
        if mansion == x:
            match += 1
            print("match:" + str(match))
    if match > 0:
        return True




def get_mansion(load_url, times):
    # データベースに接続する
    conn = sqlite3.connect('teikyou_hantei.db', detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES)
    cur = conn.cursor()
    dbname = 'teikyou_hantei.db'
    transaction_id = randomname(8)
    # "TIMESTAMP"コンバータ関数 をそのまま ”DATETIME” にも使う
    sqlite3.dbapi2.converters['DATETIME'] = sqlite3.dbapi2.converters['TIMESTAMP']

    try:
        speed = 0
        speed_han = 2

        p = pathlib.Path('../mansionProject/chromedriver')
        print(p.cwd())
        driver = webdriver.Chrome(p)

        if table_isexist(conn, cur) == False:
            create_table(conn, cur)
            db_id = 1
        else:
            try:
                db_id = cur.execute('SELECT id FROM teikyou_hantei ORDER BY id DESC LIMIT 1').fetchone()[0]
                print(db_id)
                db_id += 1
            except Exception as e:
                print(e)
                db_id = 1

        mansions = []
        mansion_url = []

        i = 1
        if times == "一括":
            if "&page" in load_url:
                load_url = re.sub('&([0-9a-zA-Z]*)=([0-9]*)', "", load_url)
                print("正規化:" + load_url)
            elif "?page" in load_url:
                load_url = re.sub('?([0-9a-zA-Z]*)=([0-9]*)', "", load_url)
                print("正規化:" + load_url)
            load_url = load_url + '?page=' + str(i)
            html = requests.get(load_url)
            soup = BeautifulSoup(html.content, "html.parser")
            print(load_url)
            city_name = soup.find(id='breadcrumb_3').text
            try:
                i_max = soup.find(class_='pagination__current-page').text
                i_max = i_max.replace("1/", "")
            except:
                i_max = 2
            print(i_max)
            while i < int(i_max):
                load_url = load_url + '?page=' + str(i)
                time.sleep(3)
                html = requests.get(load_url)
                soup = BeautifulSoup(html.content, "html.parser")
                # すべてのheadingクラスを検索して、その文字列を表示する
                for element in soup.find_all(class_="heading"):
                    try:
                        print(element.text)
                        if 'の建物' in element.text:
                            continue
                        elif db_search(cur, element.text) == True:
                            print('マンション名過去取得済み')
                            continue
                        elif past_search(mansions, element.text) == True:
                            print('マンション名が配列に重複しています')
                            continue
                        else:
                            mansions.append(element.text)
                    except Exception as e:
                        print(e)
                        mansions.append(element.text)
                        print(element.text)
                i += 1
        else:
            while i < 2:
                html = requests.get(load_url)
                soup = BeautifulSoup(html.content, "html.parser")
                print(load_url)

                city_name = soup.find(id='breadcrumb_3').text
                # すべてのheadingクラスを検索して、その文字列を表示する
                for element in soup.find_all(class_="heading"):
                    try:
                        print(element.text)
                        if 'の建物' in element.text:
                            continue
                        elif db_search(cur, element.text) == True:
                            print('マンション名過去取得済み')
                            continue
                        elif past_search(mansions, element.text) == True:
                            print('マンション名が配列に重複しています')
                            continue
                        mansions.append(element.text)
                    except Exception as e:
                        print(e)
                        mansions.append(element.text)
                        print(element.text)
                # elif site == "HOME'S":
                #     html = requests.get(load_url)
                #     soup = BeautifulSoup(html.content, "html.parser")
                #     print(load_url)
                #
                #     print(soup)
                #     city_name = soup.find(id='mod-breadcrumbs').contents[7]
                #     # すべてのheadingクラスを検索して、その文字列を表示する
                #     for element in soup.find_all(class_="bukkenName"):
                #         if '駅 徒歩' in element.text:
                #             continue
                #         mansions.append(element.text)
                #         print(element.text)
                i += 1

        print(mansions)

        driver.get("https://www.google.com/maps/")

        addresses = []
        time.sleep(4 + speed)

        element = driver.find_element_by_class_name('tactile-searchbox-input')
        element.send_keys(city_name)
        element.send_keys(Keys.ENTER)
        time.sleep(2)
        element = driver.find_element_by_id('sb_cb50')
        element.click()

        for i in range(len(mansions)):
            time.sleep(1 + speed)
            element = driver.find_element_by_class_name('tactile-searchbox-input')
            element.send_keys(mansions[i])
            element.send_keys(Keys.ENTER)
            time.sleep(2 + speed)

            try:
                element = driver.find_element_by_class_name('section-hero-header-title-subtitle')
                if element.text == "":
                    element = driver.find_element_by_class_name('ugiz4pqJLAG__primary-text')
                    if element.text == "":
                        addresses.append("取得できませんでした")
                addresses.append(element.text)
                time.sleep(2 + speed)
            except Exception as e:
                print(e)
                addresses.append("取得できませんでした")

            element = driver.find_element_by_id('sb_cb50')
            element.click()

        print(addresses)

        # ワークブックを新規作成する
        book = openpyxl.Workbook()
        # シートを取得し名前を変更する
        sheet = book.active
        sheet.title = '提供判定結果'

        sheet.cell(row=1, column=1).value = "マンション名"
        sheet.cell(row=1, column=2).value = "住所"
        sheet.cell(row=1, column=3).value = "判定結果"

        zip1 = []
        zip2 = []
        addresses_num3 = []
        addresses_num4 = []
        addresses_num5 = []

        for j in range(len(mansions)):
            print(mansions[j])
            print(addresses[j])
            sheet.cell(row=j + 2, column=1).value = mansions[j]  # セルに値を設定する
            sheet.cell(row=j + 2, column=2).value = addresses[j]

            address = addresses[j]

            if address == '取得できませんでした':
                zip1.append("")
                zip2.append("")
                addresses_num3.append("")
                addresses_num4.append("")
                addresses_num5.append("")
                continue

            try:
                post_code = re.findall(r'\d+', address)
                print(post_code)

                if len(post_code[0]) == 3 and len(post_code[1]) == 4:
                    zip1.append(post_code[0])
                    zip2.append(post_code[1])
                    addresses_num3.append(post_code[2])
                    addresses_num4.append(post_code[3])
                    try:
                        if post_code[4] != None:
                            addresses_num5.append(post_code[4])
                        else:
                            addresses_num5.append("")
                    except Exception as e:
                        print(e)
                        addresses_num5.append("")
                else:
                    print(e)
                    zip1.append("")
                    zip2.append("")
                    addresses_num3.append("")
                    addresses_num4.append("")
                    addresses_num5.append("")
            except Exception as e:
                print(e)
                zip1.append("")
                zip2.append("")
                addresses_num3.append("")
                addresses_num4.append("")
                addresses_num5.append("")


        print(zip1)
        print(zip2)
        print(addresses_num3)
        print(addresses_num4)

        # 幅調整
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 50

        driver.close()

        # ワークブックに名前をつけて保存する
        book.save('result.xlsx')

        driver = webdriver.Chrome(p)
        for z in range(len(addresses)):

            if "取得できません" in addresses[z] or zip1[z] == "":
                continue

            driver.get("https://flets-w.com/cart/?etc_data_kks=kantan%3D1%40")
            time.sleep(5 + speed_han)

            print("----------------------------")
            print(mansions[z])
            print(addresses[z])

            results = []
            driver.maximize_window()

            element = driver.find_element_by_id('p_mansion_1')
            element.click()
            time.sleep(1 + speed_han)

            element = driver.find_element_by_id('zip1')
            element.send_keys(zip1[z])
            time.sleep(1 + speed_han)

            element = driver.find_element_by_id('zip2')
            element.send_keys(zip2[z])
            time.sleep(1 + speed_han)

            element = driver.find_element_by_class_name('btn_area_zip')
            element.click()
            time.sleep(3 + speed_han)

            if '丁目' in addresses[z]:

                if addresses_num3[z] != "":
                    print("丁目：" + addresses_num3[z])
                    try:
                        elements = driver.find_elements_by_partial_link_text(addresses_num3[z])
                        print("丁目が数件部分一致")
                        element = elements[1]

                    except Exception as e:
                        print(e)
                        try:
                            print("丁目を1件再検索")
                            time.sleep(speed_han)
                            element = driver.find_element_by_partial_link_text(addresses_num3[z])
                        except Exception as e:
                            print(e)
                            result = "提供可否不明"
                            sheet.cell(row=z + 2, column=3).value = result
                            sheet.column_dimensions['C'].width = 20
                            print("id:" + str(db_id) + " マンション名:" + mansions[z] + " 住所:" + addresses[z] + " 提供結果:" + result + " トランザクションID:" + transaction_id + " 日付:" + str(datetime.datetime.now()))
                            cur.execute('INSERT INTO teikyou_hantei VALUES (?, ?, ?, ?, ?, ?)',
                                        (db_id, mansions[z], addresses[z], result, transaction_id, datetime.datetime.now()))
                            db_id += 1
                            # ワークブックに名前をつけて保存する
                            book.save('result.xlsx')
                            time.sleep(5 + speed_han)
                            continue

                    element.click()
                    time.sleep(2 + speed_han)


                if addresses_num4[z] != "":
                    print("番地：" + addresses_num4[z])
                    try:
                        element = driver.find_element_by_link_text(addresses_num4[z])
                        # if len(elements) > 1:
                        #     element = elements[1]
                        # else:
                        #     element = elements[0]
                        print("番地が一致しました")
                        element.click()
                        time.sleep(2 + speed_han)
                    except Exception as e:
                        print(e)
                        print("番地を1件再検索")
                        element = driver.find_element_by_partial_link_text(addresses_num4[z])
                        element.click()
                        time.sleep(2 + speed_han)

                if addresses_num5[z] != "":
                    print("号：" + addresses_num5[z])
                    try:
                        print("号を検索")
                        elements = driver.find_elements_by_link_text(addresses_num5[z])
                        if len(elements) > 1:
                            element = elements[1]
                        else:
                            element = elements[0]

                        element.click()
                        time.sleep(2 + speed_han)
                    except Exception as e:
                        print(e)
                        try:
                            print("号を1件再検索")
                            element = driver.find_element_by_link_text(addresses_num5[z])
                            element.click()
                        except Exception as e:
                            print(e)
                            result = "提供可否不明"
                            sheet.cell(row=z + 2, column=3).value = result
                            sheet.column_dimensions['C'].width = 20
                            print("id:" + str(db_id) + " マンション名:" + mansions[z] + " 住所:" + addresses[z] + " 提供結果:" + result + " トランザクションID:" + transaction_id + " 日付:" + str(datetime.datetime.now()))
                            cur.execute('INSERT INTO teikyou_hantei VALUES (?, ?, ?, ?, ?, ?)', (db_id, mansions[z], addresses[z], result, transaction_id, datetime.datetime.now()))
                            db_id += 1
                            # ワークブックに名前をつけて保存する
                            book.save('result.xlsx')
                            time.sleep(5 + speed_han)
                            continue

                else:
                    result = "提供可否不明"
                    sheet.cell(row=z + 2, column=3).value = result
                    sheet.column_dimensions['C'].width = 20

                    # ワークブックに名前をつけて保存する
                    book.save('result.xlsx')
                    time.sleep(5 + speed_han)
                    continue
            else:
                print("丁目なし")
                print("番地" + addresses_num3[z])
                element = driver.find_element_by_xpath('//*[@id="list-box"]/ul/li/a')
                element.click()
                time.sleep(speed_han)
                try:
                    element = driver.find_element_by_link_text(addresses_num3[z])
                except Exception as e:
                    print(e)
                    try:
                        print("丁目を1件再検索")
                        time.sleep(speed_han)
                        element = driver.find_element_by_link_text(addresses_num3[z])
                    except Exception as e:
                        print(e)
                        print("丁目がないため1つ目をクリック")
                        element = driver.find_element_by_xpath('//*[@id="list-box"]/ul/li/a')

                element.click()
                time.sleep(2 + speed_han)

                print("号：" + addresses_num4[z])
                try:
                    element = driver.find_element_by_link_text(addresses_num4[z])
                    print("号が一致しました")
                    element.click()
                    time.sleep(2 + speed_han)
                except Exception as e:
                    try:
                        print(e)
                        print("号を1件再検索")
                        element = driver.find_element_by_partial_link_text(addresses_num4[z])
                        element.click()
                        time.sleep(2 + speed_han)
                    except Exception as e:
                        print(e)
                        result = "提供可否不明"
                        sheet.cell(row=z + 2, column=3).value = result
                        sheet.column_dimensions['C'].width = 20
                        print("id:" + str(db_id) + " マンション名:" + mansions[z] + " 住所:" + addresses[z] + " 提供結果:" + result + " トランザクションID:" + transaction_id + " 日付:" + str(datetime.datetime.now()))
                        cur.execute('INSERT INTO teikyou_hantei VALUES (?, ?, ?, ?, ?, ?)', (db_id, mansions[z], addresses[z], result, transaction_id, datetime.datetime.now()))
                        db_id += 1
                        # ワークブックに名前をつけて保存する
                        book.save('result.xlsx')
                        time.sleep(5 + speed_han)
                        continue

            try:
                element = driver.find_element_by_xpath('//*[@id="list-box"]/ul/li/a')
                if element.text != "":
                    element.click()
                    time.sleep(4 + speed_han)

                hantei_address = driver.find_element_by_id('kakutei-box').text
                print(hantei_address)
                element = driver.find_element_by_xpath('//*[@id="btn-hantei"]/a')
                element.click()
                time.sleep(8 + speed_han)
            except Exception as e:
                print(e)
                try:
                    hantei_address = driver.find_element_by_id('kakutei-box').text
                    print(hantei_address)
                    element = driver.find_element_by_xpath('//*[@id="btn-hantei"]/a')
                    element.click()
                    time.sleep(8 + speed_han)
                except Exception as e:
                    print(e)
                    element = driver.find_element_by_xpath('//*[@id="list-box"]/ul/li/a')
                    element.click()

            try:
                time.sleep(3)
                element = driver.find_element_by_id('tab_wrap_inner')
                print(element.text)
            except Exception as e:
                print(e)
                element = driver.find_element_by_xpath('//*[@id="btn-hantei"]/a')
                element.click()
                time.sleep(8 + speed_han)

                element = driver.find_element_by_id('tab_wrap_inner')
                print(element.text)



            if "ファミリー・" in element.text:
                result = "ファミリー隼"
            elif "提供条件が整った場合にご提供いたしますので" in element.text:
                result = "提供不可"
            elif "お申し込み可能なサービスやご提供時期については弊社にて調査後" in element.text:
                result = "調査中"
            elif "ＶＤＳＬ方式" in element.text:
                result = "ＶＤＳＬ方式"
            elif "マンションミニ" in element.text:
                result = "ミニ隼"
            elif "マンション・スーパーハイスピードタイプ 隼" in element.text:
                result = "隼"
            else:
                result = element.text

            results.append(result)
            # sheet.cell(row=z + 2, column=3).value = hantei_address
            sheet.cell(row=z + 2, column=3).value = result
            sheet.column_dimensions['C'].width = 20
            sheet.column_dimensions['D'].width = 20

            print("id:" + str(db_id) + " マンション名:" + mansions[z] + " 住所:" + addresses[z] + " 提供結果:" + result + " トランザクションID:" + transaction_id + " 日付:" + str(datetime.datetime.now()))
            cur.execute('INSERT INTO teikyou_hantei VALUES (?, ?, ?, ?, ?, ?)', (db_id, mansions[z], addresses[z], result, transaction_id, datetime.datetime.now()))
            # 保存を実行（忘れると保存されないので注意）
            conn.commit()
            db_id += 1

            # ワークブックに名前をつけて保存する
            book.save('result.xlsx')

            time.sleep(5 + speed_han)

        driver.close()

        print("提供判定が完了しました")
        try:
            contents = cur.execute('SELECT * FROM teikyou_hantei WHERE transaction_id = ?', (transaction_id,)).fetchall()
            set_mappin.set_mappin(transaction_id)
        except Exception as e:
            print(e)
        # DB接続を閉じる
        # conn.close()
        return contents

    except Exception as e:
        print("提供判定に失敗しました")
        print(e)
        e = e
        return e
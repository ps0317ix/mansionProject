import requests
from bs4 import BeautifulSoup
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import re
import pathlib

def get_address(load_url, speed, driver):
    try:
        mansions = []
        # 取得したマンション名を格納
        # load_url = 'https://door.ac/list?utf8=%E2%9C%93&cond%5Bcities%5D%5B%5D=27103&cond%5Bsort%5D=inquiry_price&cond%5Bcities%5D%5B%5D=27103&cond%5Bfee_min%5D=&cond%5Bfee_max%5D=&cond%5Blayouts%5D%5B%5D=11&cond%5Blayouts%5D%5B%5D=12&cond%5Bwalk_time%5D=&cond%5Bsqmeter_min%5D=&cond%5Bsqmeter_max%5D=&cond%5Bage%5D='  # 初期値
        i = 1
        while i < 2:
            # load_url = load_url + '?page=' + str(i)
            html = requests.get(load_url)
            soup = BeautifulSoup(html.content, "html.parser")
            print(load_url)

            # すべてのheadingクラスを検索して、その文字列を表示する
            for element in soup.find_all(class_="heading"):
                if 'の建物' in element.text:
                    continue
                mansions.append(element.text)
                print(element.text)

            load_url = load_url
            i += 1

        print(mansions)

        driver.get("https://www.google.com/maps/")

        addresses = []
        time.sleep(4 + speed)

        for i in range(len(mansions)):
            time.sleep(1 + speed)
            element = driver.find_element_by_class_name('tactile-searchbox-input')
            element.send_keys(mansions[i])
            element.send_keys(Keys.ENTER)
            time.sleep(2 + speed)

            try:
                element = driver.find_element_by_class_name('section-hero-header-title-subtitle')
                if element.text == "":
                    element = driver.find_element_by_class_name('ugiz4pqJLAG__primary-text')
                    if element.text == "":
                        addresses.append("取得できませんでした")
                addresses.append(element.text)
                time.sleep(2 + speed)
            except:
                addresses.append("取得できませんでした")

            element = driver.find_element_by_id('sb_cb50')
            element.click()

        print(addresses)

        # ワークブックを新規作成する
        book = openpyxl.Workbook()
        # シートを取得し名前を変更する
        sheet = book.active
        sheet.title = '提供判定結果'

        sheet.cell(row=1, column=1).value = "マンション名"
        sheet.cell(row=1, column=2).value = "住所"
        sheet.cell(row=1, column=3).value = "判定結果"

        zip1 = []
        zip2 = []
        addresses_num3 = []
        addresses_num4 = []
        addresses_num5 = []

        for j in range(len(mansions)):
            print(mansions[j])
            print(addresses[j])
            sheet.cell(row=j + 2, column=1).value = mansions[j]  # セルに値を設定する
            sheet.cell(row=j + 2, column=2).value = addresses[j]

            address = addresses[j]

            if address == '取得できませんでした':
                zip1.append("")
                zip2.append("")
                addresses_num3.append("")
                addresses_num4.append("")
                addresses_num5.append("")
                continue
            post_code = re.findall(r'\d+', address)
            print(post_code)

            zip1.append(post_code[0])
            zip2.append(post_code[1])
            addresses_num3.append(post_code[2])
            addresses_num4.append(post_code[3])
            try:
                if post_code[4] != None:
                    addresses_num5.append(post_code[4])
            except:
                addresses_num5.append("")

        print(zip1)
        print(zip2)
        print(addresses_num3)
        print(addresses_num4)

        # 幅調整
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 50

        driver.close()

        # ワークブックに名前をつけて保存する
        book.save('result.xlsx')

    except:
        print("提供判定に失敗しました")
        return "提供判定に失敗しました"


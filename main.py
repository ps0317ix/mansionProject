from app.app import app
import requests
from bs4 import BeautifulSoup
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import re
import pathlib
import set_results

if __name__ == '__main__':
    app.run(debug=True)



    # try:
    # speed = 2
    # speed_han = 2
    #
    # p = pathlib.Path('../mansionProject/chromedriver')
    # print(p.cwd())
    # driver = webdriver.Chrome(p)
    #
    #
    # mansions = []
    # # 取得したマンション名を格納
    # load_url = 'https://door.ac/list?utf8=%E2%9C%93&cond%5Bcities%5D%5B%5D=27103&cond%5Bsort%5D=inquiry_price&cond%5Bcities%5D%5B%5D=27103&cond%5Bfee_min%5D=&cond%5Bfee_max%5D=&cond%5Blayouts%5D%5B%5D=11&cond%5Blayouts%5D%5B%5D=12&cond%5Bwalk_time%5D=&cond%5Bsqmeter_min%5D=&cond%5Bsqmeter_max%5D=&cond%5Bage%5D='  # 初期値


    # i = 1
    # while i < 2:
    #     # load_url = load_url + '?page=' + str(i)
    #     html = requests.get(load_url)
    #     soup = BeautifulSoup(html.content, "html.parser")
    #     print(load_url)
    #
    #     # すべてのheadingクラスを検索して、その文字列を表示する
    #     for element in soup.find_all(class_="heading"):
    #         if 'の建物' in element.text:
    #             continue
    #         mansions.append(element.text)
    #         print(element.text)
    #
    #     load_url = load_url
    #     i += 1
    #
    # print(mansions)
    #
    # driver.get("https://www.google.com/maps/")
    #
    # addresses = []
    # time.sleep(4 + speed)
    #
    # for i in range(len(mansions)):
    #     time.sleep(1 + speed)
    #     element = driver.find_element_by_class_name('tactile-searchbox-input')
    #     element.send_keys(mansions[i])
    #     element.send_keys(Keys.ENTER)
    #     time.sleep(2 + speed)
    #
    #     try:
    #         element = driver.find_element_by_class_name('section-hero-header-title-subtitle')
    #         if element.text == "":
    #             element = driver.find_element_by_class_name('ugiz4pqJLAG__primary-text')
    #             if element.text == "":
    #                 addresses.append("取得できませんでした")
    #         addresses.append(element.text)
    #         time.sleep(2 + speed)
    #     except:
    #         addresses.append("取得できませんでした")
    #
    #     element = driver.find_element_by_id('sb_cb50')
    #     element.click()
    #
    # print(addresses)
    #
    # # ワークブックを新規作成する
    # book = openpyxl.Workbook()
    # # シートを取得し名前を変更する
    # sheet = book.active
    # sheet.title = '提供判定結果'
    #
    # sheet.cell(row=1, column=1).value = "マンション名"
    # sheet.cell(row=1, column=2).value = "住所"
    # sheet.cell(row=1, column=3).value = "判定結果"
    #
    # zip1 = []
    # zip2 = []
    # addresses_num3 = []
    # addresses_num4 = []
    # addresses_num5 = []
    #
    # for j in range(len(mansions)):
    #     print(mansions[j])
    #     print(addresses[j])
    #     sheet.cell(row=j + 2, column=1).value = mansions[j]  # セルに値を設定する
    #     sheet.cell(row=j + 2, column=2).value = addresses[j]
    #
    #     address = addresses[j]
    #
    #     if address == '取得できませんでした':
    #         zip1.append("")
    #         zip2.append("")
    #         addresses_num3.append("")
    #         addresses_num4.append("")
    #         addresses_num5.append("")
    #         continue
    #     post_code = re.findall(r'\d+', address)
    #     print(post_code)
    #
    #     zip1.append(post_code[0])
    #     zip2.append(post_code[1])
    #     addresses_num3.append(post_code[2])
    #     addresses_num4.append(post_code[3])
    #     try:
    #         if post_code[4] != None:
    #             addresses_num5.append(post_code[4])
    #     except:
    #         addresses_num5.append("")
    #
    # print(zip1)
    # print(zip2)
    # print(addresses_num3)
    # print(addresses_num4)
    #
    # # 幅調整
    # sheet.column_dimensions['A'].width = 40
    # sheet.column_dimensions['B'].width = 50
    #
    # driver.close()
    #
    # # ワークブックに名前をつけて保存する
    # book.save('result.xlsx')


    # addresses = get_address.addresses
    #
    # for z in range(len(addresses)):
    #
    #     if "取得できません" in addresses[z]:
    #         continue
    #
    #     driver.get("https://flets-w.com/cart/?etc_data_kks=kantan%3D1%40")
    #     time.sleep(5 + speed_han)
    #
    #     print("----------------------------")
    #     print(addresses[z])
    #
    #     driver.maximize_window()
    #
    #     element = driver.find_element_by_id('p_mansion_1')
    #     element.click()
    #     time.sleep(1 + speed_han)
    #
    #
    #     element = driver.find_element_by_id('zip1')
    #     element.send_keys(get_address.zip1[z])
    #     time.sleep(1 + speed_han)
    #
    #     element = driver.find_element_by_id('zip2')
    #     element.send_keys(get_address.zip2[z])
    #     time.sleep(1 + speed_han)
    #
    #     element = driver.find_element_by_class_name('btn_area_zip')
    #     element.click()
    #     time.sleep(3 + speed_han)
    #
    #     addresses_num3 = get_address.addresses_num3
    #     if addresses_num3[z] != "":
    #         print("丁目：" + addresses_num3[z])
    #         try:
    #             elements = driver.find_elements_by_partial_link_text(addresses_num3[z])
    #             print("丁目が数件部分一致")
    #             element = elements[1]
    #
    #         except:
    #             try:
    #                 print("丁目を1件再検索")
    #                 time.sleep(speed_han)
    #                 element = driver.find_element_by_partial_link_text(addresses_num3[z])
    #             except:
    #                 print("丁目がないため1つ目をクリック")
    #                 element = driver.find_element_by_xpath('//*[@id="list-box"]/ul/li/a')
    #
    #         element.click()
    #         time.sleep(2 + speed_han)
    #
    #     addresses_num4 = get_address.addresses_num4
    #     if addresses_num4[z] != "":
    #         print("番地：" + addresses_num4[z])
    #         try:
    #             element = driver.find_element_by_link_text(addresses_num4[z])
    #             # if len(elements) > 1:
    #             #     element = elements[1]
    #             # else:
    #             #     element = elements[0]
    #             print("番地が一致しました")
    #             element.click()
    #             time.sleep(2 + speed_han)
    #         except:
    #             print("番地を1件再検索")
    #             element = driver.find_element_by_partial_link_text(addresses_num4[z])
    #             element.click()
    #             time.sleep(2 + speed_han)
    #
    #     addresses_num5 = get_address.addresses_num5
    #     sheet = get_address.sheet
    #     if addresses_num5[z] != "":
    #         print("号：" + addresses_num5[z])
    #         try:
    #             print("号を1件再検索")
    #             elements = driver.find_elements_by_link_text(addresses_num5[z])
    #             if len(elements) > 1:
    #                 element = elements[1]
    #                 element.click()
    #             else:
    #                 element = elements[0]
    #
    #             time.sleep(2 + speed_han)
    #         except:
    #             try:
    #                 element = driver.find_element_by_link_text(addresses_num5[z])
    #                 element.click()
    #             except:
    #                 result = "提供可否不明"
    #                 sheet.cell(row=z + 2, column=3).value = result
    #                 sheet.column_dimensions['C'].width = 20
    #
    #                 # ワークブックに名前をつけて保存する
    #                 book.save('result.xlsx')
    #                 time.sleep(5 + speed_han)
    #                 continue
    #
    #     else:
    #         result = "提供可否不明"
    #         sheet.cell(row=z + 2, column=3).value = result
    #         sheet.column_dimensions['C'].width = 20
    #
    #         # ワークブックに名前をつけて保存する
    #         book = get_address.book
    #         book.save('result.xlsx')
    #         time.sleep(5 + speed_han)
    #         continue
    #
    #
    #     try:
    #         element = driver.find_element_by_xpath('//*[@id="list-box"]/ul/li/a')
    #         if element.text != "":
    #             element.click()
    #             time.sleep(2 + speed_han)
    #             hantei_address = driver.find_element_by_id('kakutei-box').text
    #             print(hantei_address)
    #
    #         element = driver.find_element_by_xpath('//*[@id="btn-hantei"]/a')
    #         element.click()
    #         time.sleep(8 + speed_han)
    #     except:
    #         try:
    #             element = driver.find_element_by_xpath('//*[@id="btn-hantei"]/a')
    #             element.click()
    #             time.sleep(8 + speed_han)
    #         except:
    #             element = driver.find_element_by_xpath('//*[@id="list-box"]/ul/li/a')
    #             element.click()
    #
    #     try:
    #         element = driver.find_element_by_id('tab_wrap_inner')
    #         print(element.text)
    #     except:
    #         element = driver.find_element_by_xpath('//*[@id="btn-hantei"]/a')
    #         element.click()
    #         time.sleep(8 + speed_han)
    #
    #         element = driver.find_element_by_id('tab_wrap_inner')
    #         print(element.text)
    #
    #     print()
    #
    #     if "ファミリー・" in element.text:
    #         result = "ファミリー隼"
    #     elif "提供条件が整った場合にご提供いたしますので" in element.text:
    #         result = "提供不可"
    #     elif "お申し込み可能なサービスやご提供時期については弊社にて調査後" in element.text:
    #         result = "調査中"
    #     elif "ＶＤＳＬ方式" in element.text:
    #         result = "ＶＤＳＬ方式"
    #     elif "マンションミニ" in element.text:
    #         result = "ミニ隼"
    #     elif "マンション・スーパーハイスピードタイプ 隼" in element.text:
    #         result = "隼"
    #     else:
    #         result = element.text
    #
    #     # sheet.cell(row=z + 2, column=3).value = hantei_address
    #     sheet.cell(row=z + 2, column=3).value = result
    #     sheet.column_dimensions['C'].width = 20
    #     sheet.column_dimensions['D'].width = 20
    #
    #     # ワークブックに名前をつけて保存する
    #     book.save('result.xlsx')
    #
    #     time.sleep(5 + speed_han)
    #
    #     driver.close()




    #         return "提供判定が完了しました"
    # except:
    #     return "提供判定に失敗しました"